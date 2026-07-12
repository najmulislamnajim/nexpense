import csv

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

COLUMNS = ["Date", "Category", "Amount", "Payment Method", "Note"]


def _rows(queryset):
    for expense in queryset:
        yield [
            expense.date.isoformat(),
            expense.category.name,
            str(expense.amount),
            expense.get_payment_method_display(),
            expense.note,
        ]


def export_csv(queryset, filename="nexpense-report"):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    writer.writerow(COLUMNS)
    writer.writerows(_rows(queryset))
    return response


def export_xlsx(queryset, filename="nexpense-report"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    total = 0
    for row in _rows(queryset):
        ws.append(row)
        total += float(row[2])
    ws.append([])
    ws.append(["", "", "", "Total", round(total, 2)])
    ws["D" + str(ws.max_row)].font = Font(bold=True)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response


def export_pdf(queryset, filename="nexpense-report", user=None, date_range=""):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("nexpense — Expense Report", styles["Title"]),
        Paragraph(f"Generated {timezone.localdate().isoformat()} for {user}" if user else "", styles["Normal"]),
        Paragraph(date_range, styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    data = [COLUMNS] + list(_rows(queryset))
    total = sum(float(r[2]) for r in data[1:]) if len(data) > 1 else 0
    data.append(["", "", "", "Total", f"{total:.2f}"])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#059669")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8fafc")]),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return response
